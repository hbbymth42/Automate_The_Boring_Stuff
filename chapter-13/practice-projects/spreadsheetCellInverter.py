#! python3
# spreadsheetCellInverter.py - Inverts cells in spreadsheet

import sys, openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('invertedVals.xlsx')
beforeSheet = wb[wb.sheetnames[0]]
beforeSheet.title = 'before'

wb.create_sheet(title="after")
afterSheet = wb['after']

for row in beforeSheet['A1': get_column_letter(beforeSheet.max_column) + str(beforeSheet.max_row)]:
    for rowCell in row:
        afterSheet[get_column_letter(beforeSheet[rowCell.coordinate].row) + str(beforeSheet[rowCell.coordinate].column)] = beforeSheet[rowCell.coordinate].value

wb.save('invertedVals.xlsx')