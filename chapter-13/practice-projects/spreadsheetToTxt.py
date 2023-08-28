#! python3
# spreadsheetToTxt.py - Convert Spreadsheet to Text Files

import openpyxl
from pathlib import Path
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('invertedVals.xlsx')
sheet = wb['after']

for colNum in range(sheet.max_column):
    file = open(f'testFile{str(colNum+1)}.txt', "w")
    for rowNum in range(sheet.max_row):
        file.write(str(sheet[get_column_letter(colNum+1) + str(rowNum+1)].value) + '\n')
    file.close()