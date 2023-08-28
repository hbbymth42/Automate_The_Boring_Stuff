#! python3
# txtToSpreadsheet.py - Convert Text Files to Spreadsheet

import openpyxl
from pathlib import Path
from openpyxl.utils import get_column_letter

filePath = Path.cwd()
textFiles = list(filePath.glob('*.txt'))

wb = openpyxl.Workbook()

sheet = wb.active


for textFileNum, textFile in enumerate(textFiles):
    file = open(textFile)
    fileLines = file.readlines()
    for fileLineNum, fileLine in enumerate(fileLines):
        sheet[get_column_letter(textFileNum+1) + str(fileLineNum)] = fileLine

wb.save('textFilestoExcel.xlsx')