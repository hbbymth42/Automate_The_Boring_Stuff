#! python3
# excelToCSVConverter.py - Converts Excel File to CSV File

import openpyxl, csv, os, re
from openpyxl.utils import get_column_letter

for excelFile in os.listdir('.'):
    if not excelFile.endswith('.xlsx'):
        continue

    print(f'Converting {excelFile} to CSVs')

    wb = openpyxl.load_workbook(excelFile)
    wbName = re.findall('[\d\w]+',excelFile)[0]

    for sheetName in wb.sheetnames:
        sheet = wb[sheetName]
        csvFile = open(f"{wbName}_{sheet.title}.csv", 'w')
        outputWriter = csv.writer(csvFile)

        for rowNum in range(1, sheet.max_row +1):
            rowData = []
            for colNum in range(1, sheet.max_column+1):
                rowData.append(sheet[get_column_letter(colNum+1) + str(rowNum)].value)
            
            outputWriter.writerow(rowData)
        
        csvFile.close()